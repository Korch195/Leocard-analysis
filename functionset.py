import pandas as pd
import numpy as np
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
import pymysql

# викидаємо непотрібні колонки, рядки, вибираємо чи в нас csv, xlsx, і сепаратори (, чи ;)
def read_leo_data(route, filename="leocard.csv"):
    data = pd.read_csv(filename, sep=";", low_memory=False)
    leo_data = {str(route): df for route, df in data.groupby("line_name2")}[route].drop(columns=["timestamp6","carrier_name2","tariff_name4", "cp_name","medium_type", "card_uid", "card_log_num", "ticket_number4"]).reset_index(drop=True)
    leo_data['timestamp4'] = pd.to_datetime(leo_data['timestamp4'], dayfirst=True)
    leo_data.sort_values(by='timestamp4', inplace=True)
    return leo_data

# перевіряємо на наявність помилок gps
def check_stuck_gps(df, time_threshold=30):
    """
    Checks if any vehicles have been stationary at the same GPS location
    for a duration exceeding a specified time threshold.

    The function iterates through unique vehicles in the data, analyzing their
    movement timestamps and GPS locations to identify prolonged inactivity.

    Parameters:
        df (pd.DataFrame):
            A DataFrame containing data on vehicles, with the following columns:
            - 'vehicle_name': Identifier of the vehicle (str).
            - 'timestamp4': Timestamp of the vehicle's location data (datetime).
            - 'stationfrom_short_name3': Name of the station or location (str).
        time_threshold (int, optional):
            The threshold in minutes to determine if a vehicle is stationary.
            Default is 30 minutes.

    Returns:
        bool:
            True if a stationary vehicle is found exceeding the time threshold,
            else False.

    Limitations:
        - Assumes that timestamps in the 'timestamp4' column are already parsed
          as pandas datetime objects.
        - Works only with dataframes containing the specified column names.
        - Returns only for the first found instance of a stationary vehicle;
          does not report all such occurrences.
    """

    issues_found = False
    time_threshold_delta = pd.Timedelta(minutes=time_threshold)

    def calculate_time_delta(datetime_list):
        if len(datetime_list) < 2:
            raise ValueError("The list must contain at least 2 datetime objects.")
        sorted_datetimes = sorted(datetime_list)
        return sorted_datetimes[-1] - sorted_datetimes[0]

    for vehicle in df['vehicle_name'].unique():
        vehicle_data = df[df['vehicle_name'] == vehicle].sort_values(by='timestamp4')
        location = None
        times_list = []
        delta = None
        for i in range(1, len(vehicle_data)):
            time = vehicle_data.iloc[i]['timestamp4']
            location_ = vehicle_data.iloc[i]['stationfrom_short_name3']
            if location_ == location:
                times_list.append(time)
                delta = calculate_time_delta(times_list)
                if delta > time_threshold_delta:
                    issues_found = True
                    print(f"Vehicle {vehicle} has been stationary for {delta} since {times_list[0]} and until {times_list[-1]}")
                    return issues_found
            else:
                location = location_
                times_list = [time]
                delta = None
    return issues_found

# Read MicroGIZ data
def read_gps_data(filename):
    gps_data = pd.read_excel(filename)[["Geozone","Actual arrival time", "Vehicle", "Skipped stop"]]
    # filter with dropping skipped stops
    gps_data = gps_data[gps_data["Skipped stop"] == "No"]
    # convert to datetime
    gps_data['Actual arrival time'] = pd.to_datetime(gps_data['Actual arrival time'])
    # set day in format YYYY-MM-DD
    date = "2025-04-03"
    gps_data = gps_data[gps_data["Actual arrival time"].dt.date.astype(str) == date].reset_index(drop=True)
    # gps_data.rename(columns={"Actual arrival time": "timestamp4"}, inplace=True)
    return gps_data.sort_values(by='Actual arrival time').reset_index(drop=True)

# Convert MIcroGIZ data to by minute data
def make_every_minute_gps(gps_df):
    complete_time_range = pd.date_range(gps_df['Actual arrival time'].min().replace(hour=0, minute=0, second=0),
                                        gps_df['Actual arrival time'].max().replace(hour=23, minute=59, second=0),
                                        freq='min')
    full_df = pd.DataFrame({'Actual arrival time': complete_time_range})
    dct = {vehicle: df for vehicle, df in gps_df.groupby("Vehicle")}
    for vehicle, df in dct.items():
        df["Actual arrival time"] = df['Actual arrival time'].dt.round('min')
        dct[vehicle] =  pd.merge(full_df, df, how='left', on='Actual arrival time')
        dct[vehicle]["Vehicle"] = str(vehicle)
    return pd.concat(dct.values(), ignore_index=True).sort_values(by='Actual arrival time').reset_index(drop=True).drop(columns=["Skipped stop"])

# переробляємо Leocard в похвилинний датасет з 00:00 до 23:59
def convert_to_validations_per_minute(df):
    all_vehicles = pd.DataFrame()
    for vehicle in df['vehicle_name'].unique():
        vehicle_data = df[df['vehicle_name'] == vehicle].sort_values(by='timestamp4')
        line = vehicle_data['line_name2'].iloc[0]
        vehicle_data['timestamp4'] = pd.to_datetime(vehicle_data['timestamp4'], errors='coerce')
        vehicle_data['timestamp4'] = vehicle_data['timestamp4'].dt.round('min')
        vehicle_data['validations'] = 1
        vehicle_data = vehicle_data.groupby(['timestamp4', 'line_name2', 'vehicle_name', 'stationfrom_short_name3']).count().reset_index()
        complete_time_range = pd.date_range(vehicle_data['timestamp4'].min().replace(hour=0, minute=0, second=0),
                                            vehicle_data['timestamp4'].max().replace(hour=23, minute=59, second=0),
                                            freq='min')
        full_df = pd.DataFrame({'timestamp4': complete_time_range})
        vehicle_data = pd.merge(full_df, vehicle_data, on='timestamp4', how='left').fillna({'validations': 0})
        vehicle_data['vehicle_name'] = vehicle
        vehicle_data['line_name2'] = line
        all_vehicles = pd.concat([all_vehicles, vehicle_data], ignore_index=True)
        all_vehicles.sort_values(by='timestamp4', inplace=True)
    return all_vehicles

# ffill stops
def ffill_stops_properly(df_by_minute, vehicle_col = 'vehicle_name', time_col = 'timestamp4', location_col = 'stationfrom_short_name3'):
    vehicles_df = pd.DataFrame()
    for vehicle in df_by_minute[vehicle_col].unique():
        vehicle_df = df_by_minute[df_by_minute[vehicle_col] == vehicle]
        current_stop = None
        for i in range(0, len(vehicle_df[time_col])):
            if current_stop is not None:
                if str(vehicle_df[location_col].iloc[i]) == 'nan':
                    vehicle_df.loc[vehicle_df.index[i], location_col] = current_stop
                else:
                    current_stop = vehicle_df[location_col].iloc[i]
            else:
                if str(vehicle_df[location_col].iloc[i]) != 'nan':
                    current_stop = vehicle_df[location_col].iloc[i]
        vehicles_df = pd.concat([vehicles_df, vehicle_df], ignore_index=True)
    return vehicles_df

def translate_route_name(route_name):
    if isinstance(route_name, str) and route_name.startswith('ТРОЛЕЙБУС '):
        almost_numeric_part = route_name.replace('ТРОЛЕЙБУС ', '')
        almost_numeric_part = almost_numeric_part.replace('Т', '')
        numeric_part = ''.join(char for char in route_name if char.isdigit())
        if int(numeric_part) < 10:
            almost_numeric_part = '0' + almost_numeric_part
        new_route_name = 'Тр' + almost_numeric_part.lower()
        numeric_part = ''.join(char for char in route_name if char.isdigit())
    elif isinstance(route_name, str) and route_name.startswith('ТРАМВАЙ '):
        almost_numeric_part = route_name.replace('ТРАМВАЙ ', '')
        almost_numeric_part = almost_numeric_part.replace('Т', '')
        numeric_part = ''.join(char for char in route_name if char.isdigit())
        if int(numeric_part) < 10:
            almost_numeric_part = '0' + almost_numeric_part
        new_route_name = 'Т' + almost_numeric_part.lower()
    elif isinstance(route_name, str) and route_name.startswith('АВТОБУС '):
        almost_numeric_part = route_name.replace('АВТОБУС ', '')
        numeric_part = ''.join(char for char in route_name if char.isdigit())
        if int(numeric_part) < 10:
            almost_numeric_part = '0' + almost_numeric_part
        new_route_name = 'А' + almost_numeric_part.lower()
    else:
        raise Exception(f"Invalid route name: {route_name}")
    return new_route_name

# Get endpoints for this route
def get_endpoints_wrapper(df_by_min):
    """
    Extracts endpoint coordinates for a specified route from the given
    dataframe. The function ensures that the dataframe corresponds to
    a unique route, translates the route name, and retrieves its endpoints.
    If multiple routes are detected in the dataframe, an exception is raised.

    Args:
        df_by_min (DataFrame): The dataframe containing route information.
            It must include a column named 'line_name2'.
            I use the by minute dataframe(before active validations).

    Returns:
        list: A list representing the endpoints of the translated route.

    Raises:
        Exception: If more than one unique route name is found in
            the dataframe, the function raises an error indicating
            the presence of multiple routes.
    """

    def get_route_endpoints(route_name):

        def get_connection():
            connection = pymysql.connect(
                host='transport.lviv.ua',
                user='transportlviv_ucu',
                password='busbooost',
                database='transportlviv_ucu',
                charset='utf8mb4'
            )
            return connection

        def export_from_db(table):
            connection = get_connection()
            query = f"SELECT * FROM {table}"
            df = pd.read_sql(query, connection)

            connection.close()

            return df

        dtype_spec = {5: str}
        stadium = export_from_db('stadium')

        route_endpoints = stadium[(stadium['route'] == route_name) & (stadium['is_last'] == 1)]['stop_id'].unique()

        return route_endpoints

    route_name = df_by_min['line_name2'].unique()
    # print(route_name)
    if len(route_name) > 1:
        raise Exception(f"Warning: Multiple routes found in the data: {route_name}")
    new_route_name = translate_route_name(route_name[0])
    # print(new_route_name)
    route_endpoints = get_route_endpoints(new_route_name).tolist()
    print(route_endpoints)
    return route_endpoints

# Replace leocard gps with MicroGIZ gps
def merge_gps_and_validations(gps_df, leo_df):
    """
        Replaces GPS from Leocard validators with MicroGIZ GPS data.

    Args:
       gps_df (DataFrame): MicroGiz dataframe containing vehicle gps data.
            It must include a column named 'Geozone'.
            It must include a column named 'Actual arrival time'.
            It must include a column named 'Vehicle'.
            It must include a column named 'Route'.
        leo_df (DataFrame): LeoCard dataframe containing validation data.
            It must include a column named 'stationfrom_short_name3'.
            It must include a column named 'timestamp4'.
            It must include a column named 'vehicle_name'.
            It must include a column named 'line_name2'.

    Returns:
        df: with the values of the 'stationfrom_short_name3' column replased by the gps data from gps_df.
        With the route, vehicle and time in the same row matched.
    """
    x = leo_df.copy()
    y = gps_df.copy()
    for i in range(0, len(y)):
        try:
            stop_num = extract_stop_number(y['Geozone'].iloc[i])
            y.loc[y.index[i], 'Geozone'] = stop_num
        except:
            # print(f"Warning: This row is broken \n{y.iloc[i]}")
            continue
    x['timestamp4'] = pd.to_datetime(x['timestamp4'])
    y['Actual arrival time'] = pd.to_datetime(y['Actual arrival time'])
    merged = x.merge(
        y,
        left_on=['vehicle_name', 'timestamp4'],
        right_on=['Vehicle', 'Actual arrival time'],
        how='left',
        suffixes=('', '_y')
    )
    mask = merged["stationfrom_short_name3"].notna() & merged["Geozone"].notna()
    x.loc[mask, "stationfrom_short_name3"] = merged.loc[mask, "Geozone"]
    return x

# convert to active validations
def check_exact_patterns(text, patterns):
    for pattern in patterns:
        regex = rf'(?<!\d){pattern}(?![\d\-\.])'

        if re.search(regex, text):
            return True
    return False

def activate(df, endpoints, validation_window=30):
    time_window_delta = pd.Timedelta(minutes=validation_window)
    # all_vehicles = pd.DataFrame()
    values = []
    for vehicle in df['vehicle_name'].unique():
        if str(vehicle) == 'nan':
            continue
        vehicle_data = df[df['vehicle_name'] == vehicle].copy()
        vehicle_data = vehicle_data.sort_values(by='timestamp4').copy()
        k = 0
        start_window = vehicle_data['timestamp4'].iloc[k]
        for i in range(len(vehicle_data['timestamp4'])):
            end_window = vehicle_data['timestamp4'].iloc[i]
            while end_window-start_window >= time_window_delta:
                k = k+1
                start_window = vehicle_data['timestamp4'].iloc[k]
            if check_exact_patterns(str(vehicle_data['stationfrom_short_name3'].iloc[i]), endpoints):
                try:
                    if not check_exact_patterns(str(vehicle_data['stationfrom_short_name3'].iloc[i-1]), endpoints):
                        k = i
                        start_window = vehicle_data['timestamp4'].iloc[k]
                except:
                    k = i
                    start_window = vehicle_data['timestamp4'].iloc[k]
            window_validations = vehicle_data[
                    (vehicle_data['timestamp4'] >= start_window) &
                    (vehicle_data['timestamp4'] <= end_window)
                ]['validations'].sum()
            # print(window_validations)\
            values.append(window_validations)
        #     vehicle_data.at[i, 'active_validations'] = window_validations
        #     vehicle_data.iloc[i,'active_validations'] = window_validations
        # all_vehicles = pd.concat([all_vehicles, vehicle_data], ignore_index=True)
    df['active_validations'] = values
    return df


# переробляємо в одну таблицю для всього маршруту в цей день (значення окремих автобусів на маршруті сумуємо)
def create_location_pivot(df):
    """
    Transform a DataFrame with timestamp, vehicle_name, location and active_validation columns
    into a pivoted DataFrame with:
    - Each row representing a minute timestamp (preserving minute frequency)
    - Each column representing a unique location
    - Cell values containing the sum of active_validations across all vehicles
      at that timestamp and location

    Parameters:
    -----------
    df : pandas.DataFrame
        DataFrame with columns including 'timestamp4', 'stationfrom_short_name3',
        'vehicle_name', and 'active_validations'

    Returns:
    --------
    pandas.DataFrame
        Pivoted DataFrame with minute-frequency timestamps as index and locations as columns
    """
    df = df.copy()

    grouped = df.groupby(['timestamp4', 'stationfrom_short_name3'])['active_validations'].sum().unstack(fill_value=0)
    full_index = pd.date_range(start='2025-04-03 00:00:00', end='2025-04-03 23:59:00', freq='1min')
    grouped = grouped.reindex(full_index, fill_value=0)
    grouped = grouped.astype(int)
    grouped['timestamp4'] = grouped.index
    grouped = grouped.reset_index(drop=True).set_index('timestamp4')
    # print(grouped.head())
    grouped.columns.name = None
    return grouped


# беремо середнє по ненульових значення і переробляєио датасет в півгодинний
def convert_to_half_hour(df, type = 'mean'):
    df.replace(0, pd.NA, inplace=True)
    if type == 'mean':
        df = df.resample('30min').mean()#.reset_index()
    elif type == 'median':
        df = df.resample('30min').median()
    elif type == 'sum':
        df = df.resample('30min').sum()
    df.fillna(0, inplace=True)
    round(df, 2)
    return df

# add stop names
def get_stop_names():
    def get_connection():
        connection = pymysql.connect(
            host='transport.lviv.ua',
            user='transportlviv_ucu',
            password='busbooost',
            database='transportlviv_ucu',
            charset='utf8mb4'
        )
        return connection

    def export_from_db(table):
        connection = get_connection()
        query = f"SELECT * FROM {table}"
        df = pd.read_sql(query, connection)
        connection.close()
        return df

    stops_df = export_from_db('stops')

    return stops_df

# робимо colormap
def color_map_df(df, stops_df):
    # Convert both column names and stop IDs to strings
    df.columns = df.columns.astype(str)
    stops_df['stop_id'] = stops_df['stop_id'].astype(str)

    # Create mapping from stop_id to stop_name
    mapping_dict = dict(zip(stops_df['stop_id'], stops_df['stop_name']))

    # Generate label row: match stop_id to stop_name
    label_row = [mapping_dict.get(col, col) for col in df.columns]

    # Create a workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Colored Data"

    # Write headers (timestamp + stop_ids)
    ws.append(["timestamp"] + list(df.columns))

    # Write label row (timestamp column is empty)
    ws.append([""] + label_row)

    # Write data rows
    for timestamp, row in df.iterrows():
        ws.append([str(timestamp)] + list(row.values))

    # Apply color formatting like Excel's default
    data_start_row = 3
    data_end_row = data_start_row + len(df) - 1
    data_cols = df.shape[1]

    for i in range(2, data_cols + 2):  # Skip timestamp
        col_letter = get_column_letter(i)
        cell_range = f"{col_letter}{data_start_row}:{col_letter}{data_end_row}"
        rule = ColorScaleRule(
            start_type='min', start_color='63BE7B',
            mid_type='percentile', mid_value=50, mid_color='FFEB84',
            end_type='max', end_color='F8696B'
        )
        ws.conditional_formatting.add(cell_range, rule)

    # Auto-adjust column widths
    for col_idx, col in enumerate(ws.iter_cols(), 1):
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    return wb

# get stadium (order of stops)
def get_route_stadium(route_name):

    def get_connection():
        connection = pymysql.connect(
            host='transport.lviv.ua',
            user='transportlviv_ucu',
            password='busbooost',
            database='transportlviv_ucu',
            charset='utf8mb4'
        )
        return connection

    def export_from_db(table):
        connection = get_connection()
        query = f"SELECT * FROM {table}"
        df = pd.read_sql(query, connection)

        connection.close()

        return df

    dtype_spec = {5: str}
    stadium = export_from_db('stadium')
    try:
        new_route = translate_route_name(route_name[0])
    except:
        try:
            new_route = translate_route_name(route_name)
        except Exception as e:
            print(e)
            return None
    routeplan = stadium[stadium['route'] == new_route]['stop_id'].tolist()

    # print(routeplan)
    good_routeplan = []
    for i in range(0, len(routeplan)-1):
        if routeplan[i] != routeplan[i+1]:
            good_routeplan.append(routeplan[i])
    # good_routeplan.append(routeplan[len(routeplan)-1])
    return good_routeplan

# sort columns
def extract_stop_number(name):
    match = re.search(r'\((\d+)', name)
    return int(match.group(1)) if match else None

def reorder_columns_by_stop_numbers(df, ordered_stop_numbers):
    new_df = pd.DataFrame()
    new_df.index = df.index
    for i in ordered_stop_numbers:
        col = df[str(i)]
        new_df[str(i)] = col

    return new_df

#
def ensure_that_stops_are_numeric(df):
    for i in range(0, len(df['stationfrom_short_name3'])):
        try:
            if str(df.iloc[i]['stationfrom_short_name3']) == 'nan':
                df.at[i, 'stationfrom_short_name3'] = np.nan
            elif str(df.iloc[i]['stationfrom_short_name3']) != 'nan':
                df.at[i, 'stationfrom_short_name3'] = extract_stop_number(df.iloc[i]['stationfrom_short_name3'])
        except:
            continue
    df['stationfrom_short_name3'] = df['stationfrom_short_name3'].astype('string')
    return df

# wrapper

def leocard_to_passenger_demand(route, save_path=None, leocard_filename="leocard.csv", microgiz_filename="microgiz.xlsx"):
    df = read_leo_data(route, leocard_filename)
    # take one day function
    date = df['timestamp4'][0].date()
    are_problems_with_gps = False #check_stuck_gps(df)
    stadium = get_route_stadium(route)
    df = convert_to_validations_per_minute(df)
    df = ffill_stops_properly(df, vehicle_col='vehicle_name', time_col='timestamp4',
                              location_col='stationfrom_short_name3')
    if are_problems_with_gps:
        gps_df = read_gps_data(microgiz_filename)
        gps_df = make_every_minute_gps(gps_df)
        gps_df = ffill_stops_properly(gps_df, vehicle_col='Vehicle', time_col='Actual arrival time',
                                      location_col='Geozone')
        df = merge_gps_and_validations(gps_df, df)
    endpoints = get_endpoints_wrapper(df)
    df = activate(df, endpoints, validation_window=25)
    df = ensure_that_stops_are_numeric(df)
    df = create_location_pivot(df)
    df.fillna(0, inplace=True)
    df = convert_to_half_hour(df)
    df = reorder_columns_by_stop_numbers(df, stadium)
    # take ceil of all numeric values
    wb = color_map_df(df, get_stop_names())
    if save_path is None:
        save_path = f"passenger_demand_{translate_route_name(route)}_{date}.xlsx"
    wb.save(save_path)

import warnings
warnings.filterwarnings("ignore")

leocard_to_passenger_demand("ТРАМВАЙ 6Т", leocard_filename="C:\\Users\\38095\Desktop\\БІЗНЕС АНАЛІЗ ТРАСПОРТ УКУ\\CurrentFinalFolder\\leocard.csv", microgiz_filename="C:\\Users\\38095\\Desktop\\БІЗНЕС АНАЛІЗ ТРАСПОРТ УКУ\\CurrentFinalFolder\\microgiz.xlsx")